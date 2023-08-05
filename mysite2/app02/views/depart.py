from django.shortcuts import render, HttpResponse, redirect
from app02.models import Department, UserInfo, PrettyNum
from app02.utils.pagination import Pagination
from app02.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm


# Create your views here.

def depart_list(request):
    """部门列表"""
    queryset = Department.objects.all()
    # print(queryset)
    return render(request, "depart_list.html", {"queryset": queryset})


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, "depart_add.html")
    Department.objects.create(title=request.POST.get('title'))
    return redirect("/depart/list")


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get('nid')
    Department.objects.filter(id=nid).delete()
    return redirect("/depart/list")


def depart_edit(request, nid):
    """编辑部门"""
    if request.method == "GET":
        depart = Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"depart": depart.title})
    Department.objects.filter(id=nid).update(title=request.POST.get("title"))
    return redirect("/depart/list")


def depart_multi(request):
    """批量上传（Excel文件）"""
    from django.core.files.uploadedfile import InMemoryUploadedFile
    from openpyxl import load_workbook
    # 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 对象传递给openpyxl
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        print(row[0].value)
        exists = Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=row[0].value)

    return redirect("/depart/list")
